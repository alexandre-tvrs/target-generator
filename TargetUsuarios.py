from datetime import timedelta
import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Side, NamedStyle, Font, PatternFill, Border, Alignment
from AD import *


def criar_estruturacao_target_usuario(nome_cliente, diretorio, qnty_dias):
    target = openpyxl.Workbook()
    guia_ativa = target.active
    guia_ativa.title = 'AD Info Asset Glossary'
    target.create_sheet('1.Cleaned Up AD List (In Scope)')
    target.create_sheet('2. Duplicate Users')
    target.create_sheet('3. Disabled Users')
    target.create_sheet('4. >' + str(qnty_dias) + ' Days Last Active')
    target.create_sheet('5. Raw Data')

    target.save(str(diretorio) + '/' + str(nome_cliente) + '-Users.xlsx')


def inserir_dados_usuario(dados_usuario, qnty_dias, diretorio, nome_do_cliente):
    caminho = (str(diretorio) + '/' + str(nome_do_cliente) + '-Users.xlsx')

    info_asset = pd.read_excel(caminho, sheet_name='AD Info Asset Glossary')

    last_day = dados_usuario['Last logon date'].max()
    cleaned_ad = dados_usuario.loc[(dados_usuario['Disabled'] == 'False') &
                                   (last_day - timedelta(days=int(qnty_dias)) <= dados_usuario['Last logon date'])]
    duplicate = dados_usuario.loc[dados_usuario['Name'].duplicated()]
    disabled_users = dados_usuario.loc[dados_usuario['Disabled'] == 'True']
    sixty_days = dados_usuario.loc[last_day - timedelta(days=int(qnty_dias)) >= dados_usuario['Last logon date']]
    raw_data = dados_usuario

    writer = pd.ExcelWriter(caminho, engine='xlsxwriter')

    info_asset.to_excel(writer, sheet_name='AD Info Asset Glossary', index=False)
    cleaned_ad.to_excel(writer, sheet_name='1.Cleaned Up AD List (In Scope)', index=False)
    duplicate.to_excel(writer, sheet_name='2. Duplicate Users', index=False)
    disabled_users.to_excel(writer, sheet_name='3. Disabled Users', index=False)
    sixty_days.to_excel(writer, sheet_name='4. >' + str(qnty_dias) + ' Days Last Active', index=False)
    raw_data.to_excel(writer, sheet_name='5. Raw Data', index=False)

    writer.save()


def finalizar_planilha_usuario(diretorio, nome_cliente, qnty_dias):
    caminho = (str(diretorio) + '/' + str(nome_cliente) + '-Users.xlsx')
    border_top = Side(border_style='dashed', color='000000')
    border_right = Side(border_style='dashed', color='000000')
    border_left = Side(border_style='dashed', color='000000')
    border_bottom = Side(border_style='dashed', color='000000')

    border_h_top = Side(border_style='thin', color='FFFFFF')
    border_h_right = Side(border_style='thin', color='FFFFFF')
    border_h_left = Side(border_style='thin', color='FFFFFF')
    border_h_bottom = Side(border_style='thin', color='FFFFFF')

    border_alert_top = Side(border_style='thick', color='000000')
    border_alert_right = Side(border_style='thick', color='000000')
    border_alert_left = Side(border_style='thick', color='000000')
    border_alert_bottom = Side(border_style='thick', color='000000')

    border_ad_info_top = Side(border_style='thin', color='000000')
    border_ad_info_right = Side(border_style='thin', color='000000')
    border_ad_info_left = Side(border_style='thin', color='000000')
    border_ad_info_bottom = Side(border_style='thin', color='000000')

    alert_style = NamedStyle('alert_style')
    alert_style.font = Font('Montserrat', size=12, bold=True, color='000000')
    alert_style.fill = PatternFill('solid', fgColor="FFFFFF")
    alert_style.border = Border(top=border_alert_top, bottom=border_alert_bottom, left=border_alert_left,
                                right=border_alert_right)
    alert_style.alignment = Alignment(horizontal='center', vertical='center')

    target_style = NamedStyle('target_style')
    target_style.font = Font('Montserrat', size=10, bold=False, color='000000')
    target_style.fill = PatternFill('solid', fgColor="FFFFFF")
    target_style.border = Border(top=border_top, bottom=border_bottom, left=border_left, right=border_right)

    target_date_style = NamedStyle('target_date_style')
    target_date_style.font = Font('Montserrat', size=10, bold=False, color='000000')
    target_date_style.fill = PatternFill('solid', fgColor="FFFFFF")
    target_date_style.border = Border(top=border_top, bottom=border_bottom, left=border_left, right=border_right)
    target_date_style.number_format = 'D/MM/YYYY'

    header_style = NamedStyle('header_style')
    header_style.font = Font('Montserrat', size=11, bold=True, color='FFFFFF')
    header_style.fill = PatternFill('solid', fgColor="D21034")
    header_style.alignment = Alignment(horizontal='center', vertical='center')
    header_style.border = Border(top=border_h_top, bottom=border_h_bottom, left=border_h_left, right=border_h_right)

    ad_info_coluna_A = NamedStyle('ad_info_coluna_A')
    ad_info_coluna_A.font = Font('Montserrat', size=10, bold=True, color='FFFFFF')
    ad_info_coluna_A.fill = PatternFill('solid', fgColor="808080")
    ad_info_coluna_A.alignment = Alignment(horizontal='left', vertical='bottom')
    ad_info_coluna_A.border = Border(top=border_ad_info_top, bottom=border_ad_info_bottom,
                                     left=border_ad_info_left, right=border_ad_info_right)

    ad_info_coluna_B = NamedStyle('ad_info_coluna_B')
    ad_info_coluna_B.font = Font('Montserrat', size=10, bold=False, color='000000')
    ad_info_coluna_B.fill = PatternFill('solid', fgColor="bdbec0")
    ad_info_coluna_B.alignment = Alignment(horizontal='left', vertical='justify')
    ad_info_coluna_B.border = Border(top=border_ad_info_top, bottom=border_ad_info_bottom,
                                     left=border_ad_info_left, right=border_ad_info_right)

    relatorio_usuarios = openpyxl.load_workbook(caminho)

    guias = relatorio_usuarios.sheetnames

    relatorio_usuarios[guias[1]].sheet_properties.tabColor = '5c5e61'
    relatorio_usuarios[guias[2]].sheet_properties.tabColor = 'd21034'
    relatorio_usuarios[guias[3]].sheet_properties.tabColor = '5c5e61'
    relatorio_usuarios[guias[4]].sheet_properties.tabColor = '5c5e61'
    relatorio_usuarios[guias[5]].sheet_properties.tabColor = 'd21034'

    coluna = ''

    # AD INFO ASSET GLOSSARY

    logo_a = openpyxl.drawing.image.Image(f'img/logo_swo_small.png')
    logo_b = openpyxl.drawing.image.Image(f'img/logo_swo_small.png')

    relatorio_usuarios[guias[0]].add_image(logo_a, 'A1')
    relatorio_usuarios[guias[0]].add_image(logo_b, 'A9')

    relatorio_usuarios[guias[0]].sheet_view.showGridLines = False
    altura_linha_um = relatorio_usuarios[guias[0]].row_dimensions[1]
    altura_linha_doze = relatorio_usuarios[guias[0]].row_dimensions[9]
    altura_linha_um.height = 36
    altura_linha_doze.height = 36
    relatorio_usuarios[guias[0]].cell(row=1, column=1).style = header_style
    relatorio_usuarios[guias[0]].cell(row=1, column=2).style = header_style
    relatorio_usuarios[guias[0]].cell(row=1, column=2).value = 'AD Info Asset - Glossary of Terms'
    relatorio_usuarios[guias[0]].cell(row=9, column=1).style = header_style
    relatorio_usuarios[guias[0]].cell(row=9, column=2).style = header_style
    relatorio_usuarios[guias[0]].cell(row=9, column=2).value = 'Definition of Headers'
    relatorio_usuarios[guias[0]].column_dimensions['A'].width = 45
    relatorio_usuarios[guias[0]].column_dimensions['B'].width = 180

    # Inserindo texto
    cleaned = relatorio_usuarios[guias[0]].cell(row=2, column=1)
    duplicate = relatorio_usuarios[guias[0]].cell(row=3, column=1)
    disabled = relatorio_usuarios[guias[0]].cell(row=4, column=1)
    last = relatorio_usuarios[guias[0]].cell(row=5, column=1)
    raw = relatorio_usuarios[guias[0]].cell(row=6, column=1)

    cleaned.value = '1. Cleaned Up AD List (In Scope)'
    duplicate.value = '2. Duplicate Users'
    disabled.value = '3. Disabled Users'
    last.value = '4. >' + str(qnty_dias) + ' Days Last Active'
    raw.value = '5. Raw Data'

    cleaned = relatorio_usuarios[guias[0]].cell(row=2, column=2)
    duplicate = relatorio_usuarios[guias[0]].cell(row=3, column=2)
    disabled = relatorio_usuarios[guias[0]].cell(row=4, column=2)
    last = relatorio_usuarios[guias[0]].cell(row=5, column=2)
    raw = relatorio_usuarios[guias[0]].cell(row=6, column=2)

    cleaned.value = "From the Raw Data, SoftwareONE scrubs the data and the devices in this tab are the " \
                    "active assets deployed in the customer's environment which are in scope within " \
                    "60 days. Anything older than 60 days is considered out of scope / decommissioned " \
                    "/ Retired (see Tab 8. >60 Days Last Active)"
    duplicate.value = 'These are devices that have reported more than once. These devices need to be ' \
                      'excluded so as to represent a true count of the actual number of unique devices.'
    disabled.value = 'These are users that have been disabled by the system administrator. ' \
                     'These users are considered as no longer running production workloads ' \
                     'and have been removed by the system administrator from the network. ' \
                     'Disabled users within last 60 days are excluded and considered out of scope.'
    last.value = 'These are the users that are older than 60 days and are considered out of scope ' \
                 '/ decommissioned / Retired.'
    raw.value = 'This is the Raw original data output before any analysis has been done. ' \
                'Data kept for future reference'

    name = relatorio_usuarios[guias[0]].cell(row=10, column=1)
    create = relatorio_usuarios[guias[0]].cell(row=11, column=1)
    disabled = relatorio_usuarios[guias[0]].cell(row=12, column=1)
    display = relatorio_usuarios[guias[0]].cell(row=13, column=1)
    email = relatorio_usuarios[guias[0]].cell(row=14, column=1)
    first = relatorio_usuarios[guias[0]].cell(row=15, column=1)
    last = relatorio_usuarios[guias[0]].cell(row=16, column=1)
    parent = relatorio_usuarios[guias[0]].cell(row=17, column=1)

    name.value = 'Name'
    create.value = 'Creation Date'
    disabled.value = 'Disabled'
    display.value = 'Display Name'
    email.value = 'Email Address'
    first.value = 'First Name'
    last.value = 'Last Logon Date'
    parent.value = 'Parent Container'

    name = relatorio_usuarios[guias[0]].cell(row=10, column=2)
    create = relatorio_usuarios[guias[0]].cell(row=11, column=2)
    disabled = relatorio_usuarios[guias[0]].cell(row=12, column=2)
    display = relatorio_usuarios[guias[0]].cell(row=13, column=2)
    email = relatorio_usuarios[guias[0]].cell(row=14, column=2)
    first = relatorio_usuarios[guias[0]].cell(row=15, column=2)
    last = relatorio_usuarios[guias[0]].cell(row=16, column=2)
    parent = relatorio_usuarios[guias[0]].cell(row=17, column=2)

    name.value = 'The name of the device'
    duplicate.value = 'Duplicated Users'
    create.value = 'Date of Users account creation'
    disabled.value = 'Disabled Users'
    display.value = "User's Display Name"
    email.value = "User's email"
    first.value = "User's first name"
    last.value = "User's last name"
    parent.value = "User's parent Container "

    for linha in range(2, 7):
        relatorio_usuarios[guias[0]].cell(row=linha, column=1).style = ad_info_coluna_A
        relatorio_usuarios[guias[0]].cell(row=linha, column=2).style = ad_info_coluna_B

    for linha in range(10, 19):
        relatorio_usuarios[guias[0]].cell(row=linha, column=1).style = ad_info_coluna_A
        relatorio_usuarios[guias[0]].cell(row=linha, column=2).style = ad_info_coluna_B

    for i in range(1, 6):
        numero_colunas = get_numero_colunas(i, caminho)
        numero_linhas = get_numero_linhas(i, caminho)
        verify = relatorio_usuarios[guias[i]]['A2'].value
        for j in range(1, numero_colunas + 1):
            if j == 1:
                coluna = 'A'
            if j == 2:
                coluna = 'B'
            if j == 3:
                coluna = 'C'
            if j == 4:
                coluna = 'D'
            if j == 5:
                coluna = 'E'
            if j == 6:
                coluna = 'F'
            if j == 7:
                coluna = 'G'
            if j == 8:
                coluna = 'H'
            if j == 9:
                coluna = 'I'
            for k in range(1, numero_linhas + 2):
                if verify is None:
                    relatorio_usuarios[guias[i]].merge_cells('A2:C5')
                    celula = relatorio_usuarios[guias[i]].cell(row=2, column=1)
                    celula.value = 'NO DATA FOUND'
                    celula.style = alert_style
                if (j == 2) | (j == 7):
                    relatorio_usuarios[guias[i]].cell(row=k, column=j).style = target_date_style
                else:
                    relatorio_usuarios[guias[i]].cell(row=k, column=j).style = target_style

            relatorio_usuarios[guias[i]].cell(row=1, column=j).style = header_style
            relatorio_usuarios[guias[i]].column_dimensions[coluna].width = 40

        relatorio_usuarios[guias[i]].auto_filter.ref = relatorio_usuarios[guias[i]].dimensions
        relatorio_usuarios[guias[i]].sheet_view.showGridLines = False
        altura = relatorio_usuarios[guias[i]].row_dimensions[1]
        altura.height = 36

    relatorio_usuarios.save(caminho)
