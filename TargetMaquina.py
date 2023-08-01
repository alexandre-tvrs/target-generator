from datetime import timedelta
import openpyxl
import pandas as pd
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Side, NamedStyle, Font, PatternFill, Border, Alignment
from AD import *


def criar_estruturacao_target_maquina(nome_cliente, diretorio, qnty_dias):
    target = openpyxl.Workbook()
    guia_ativa = target.active
    guia_ativa.title = 'AD Info Asset Glossary'
    target.create_sheet('Graphical Analysis')
    target.create_sheet('Pivot Quantity of Machines')
    target.create_sheet('Pivot InScope Devices')
    target.create_sheet('Pivot DTA Unknown')
    target.create_sheet('Pivot Types of OfS')
    target.create_sheet('1.Cleaned Up AD List (In Scope)')
    target.create_sheet('2. Duplicate Computers')
    target.create_sheet('3. Non MS OS')
    target.create_sheet('4. Unknown OS (In Scope)')
    target.create_sheet('5. DTA Devices (In Scope)')
    target.create_sheet('6. Failover Clusters')
    target.create_sheet('7. Disabled Devices')
    target.create_sheet('8. >' + str(qnty_dias) + ' Days Last Active')
    target.create_sheet('9. Raw Data')

    target.save(str(diretorio) + '/' + str(nome_cliente) + '-Machines.xlsx')


def inserir_dados_maquina(dados_maquina, qnty_dias, diretorio, nome_do_cliente):
    caminho = (str(diretorio) + '/' + str(nome_do_cliente) + '-Machines.xlsx')

    info_asset = pd.read_excel(caminho, sheet_name='AD Info Asset Glossary')
    analysis = pd.read_excel(caminho, sheet_name='Graphical Analysis')
    pvtDTA = pd.read_excel(caminho, sheet_name='Pivot DTA Unknown')
    pvtToOfs = pd.read_excel(caminho, sheet_name='Pivot Types of OfS')

    last_day = dados_maquina['Last logon date'].max()

    cleaned_ad = dados_maquina.loc[(dados_maquina['Disabled'] == 'False') &
                                   (last_day - timedelta(days=int(qnty_dias)) <= dados_maquina['Last logon date']) &
                                   ((dados_maquina['Type of device'] == 'Client') |
                                    (dados_maquina['Type of device'] == 'Server')) &
                                   dados_maquina['Operating System'].str.contains('win|Win|Windows|'
                                                                                  'SQL|System|Sharepoint')]

    duplicate = dados_maquina.loc[dados_maquina['Name'].duplicated()]

    nom_ms = dados_maquina.loc[~((dados_maquina['Operating System'].str.contains('Windows|SQL|System|Sharepoint') ==
                               True) | (dados_maquina['Operating System'].isnull()) | (
                                    dados_maquina['Operating System'] == ' ') | (
                                    dados_maquina['Operating System'] == '') |
                                 (last_day - timedelta(days=int(qnty_dias)) <= dados_maquina['Last logon date']))]

    unknown_os = dados_maquina.loc[(dados_maquina['Operating System'] == 'unknown') |
                                   (dados_maquina['Operating System'].isnull()) | (
                                    dados_maquina['Operating System'] == ' ') | (
                                    dados_maquina['Operating System'] == '') &
                                   (dados_maquina['Disabled'] == 'False') &
                                   (last_day - timedelta(days=int(qnty_dias)) <= dados_maquina['Last logon date'])]

    dta_devices = dados_maquina.loc[((dados_maquina['Parent Container'].str.contains('dev|DEV|Dev|test|Test|TEST|'
                                                                                     'patch|Patch|TST')
                                      == True) &
                                     (dados_maquina['Disabled'] == 'False') &
                                     (last_day - timedelta(days=int(qnty_dias)) <= dados_maquina['Last logon date']))]

    failover_clusters = dados_maquina.loc[(dados_maquina['Parent Container'].str.contains('cluster|Cluster|CLUSTER'))]

    disabled_devices = dados_maquina.loc[dados_maquina['Disabled'] == 'True']

    sixty_days = dados_maquina.loc[(last_day - timedelta(days=int(qnty_dias)) > dados_maquina['Last logon date']) &
                                   (dados_maquina['Disabled'] == 'False')]

    raw_data = dados_maquina

    dta = dta_devices.copy()
    dta['Type of device'] = 'DTA'
    unk = unknown_os.copy()
    unk['Type of device'] = 'unknown'
    DTA_Unk = pd.merge(dta, unk, how='outer')
    dup = duplicate.copy()
    dup['Type of device'] = 'Duplicate Computers'
    nms = nom_ms.copy()
    nms['Type of device'] = 'Non-MS OS'
    disabled = disabled_devices.copy()
    disabled['Type of device'] = 'Disabled Device'
    days = sixty_days.copy()
    days['Type of device'] = '>' + str(qnty_dias) + ' Days Last Active'
    merge1 = pd.merge(dup, nms, how='outer')
    merge2 = pd.merge(disabled, days, how='outer')
    OutScope = pd.merge(merge1, merge2, how='outer')
    oScope = OutScope
    pvtInD = pd.pivot_table(cleaned_ad, index=['Type of device'], values=['Name'], aggfunc='count')
    pvtDTA = pd.pivot_table(DTA_Unk, index=['Type of device'], values=['Name'], aggfunc='count')
    pvtToOfs = pd.pivot_table(OutScope, index=['Type of device'], values=['Name'], aggfunc='count', fill_value=0,
                              dropna=False)
    oScope['Type of device'] = 'Out of Scope'
    inScope = pd.merge(cleaned_ad, unk, how='outer')
    inScope['Type of device'] = 'In Scope'
    allMachines = pd.merge(inScope, OutScope, how='outer')
    pvtQoM = pd.pivot_table(allMachines, index=['Type of device'], values=['Name'], aggfunc='count')

    pvtQoM.rename(columns={'Name': 'Quantity of Machines'}, inplace=True)
    pvtInD.rename(columns={'Name': 'In-Scope Devices'}, inplace=True)

    writer = pd.ExcelWriter(caminho, engine='xlsxwriter')

    info_asset.to_excel(writer, sheet_name='AD Info Asset Glossary', index=False)
    analysis.to_excel(writer, sheet_name='Graphical Analysis', index=False)
    pvtQoM.to_excel(writer, sheet_name='Pivot Quantity of Machines', index=True)
    pvtInD.to_excel(writer, sheet_name='Pivot InScope Devices', index=True)
    pvtDTA.to_excel(writer, sheet_name='Pivot DTA Unknown', index=True)
    pvtToOfs.to_excel(writer, sheet_name='Pivot Types of OfS', index=True)
    cleaned_ad.to_excel(writer, sheet_name='1.Cleaned Up AD List (In Scope)', index=False)
    duplicate.to_excel(writer, sheet_name='2. Duplicate Computers', index=False)
    nom_ms.to_excel(writer, sheet_name='3. Non MS OS', index=False)
    unknown_os.to_excel(writer, sheet_name='4. Unknown OS (In Scope)', index=False)
    dta_devices.to_excel(writer, sheet_name='5. DTA Devices (In Scope)', index=False)
    failover_clusters.to_excel(writer, sheet_name='6. Failover Clusters', index=False)
    disabled_devices.to_excel(writer, sheet_name='7. Disabled Devices', index=False)
    sixty_days.to_excel(writer, sheet_name='8. >' + str(qnty_dias) + ' Days Last Active', index=False)
    raw_data.to_excel(writer, sheet_name='9. Raw Data', index=False)

    writer.save()


def finalizar_planilha_maquina(diretorio, nome_do_cliente, qnty_dias):
    caminho = (str(diretorio) + '/' + str(nome_do_cliente) + '-Machines.xlsx')
    border_top = Side(border_style='dashed', color='000000')
    border_right = Side(border_style='dashed', color='000000')
    border_left = Side(border_style='dashed', color='000000')
    border_bottom = Side(border_style='dashed', color='000000')

    border_h_top = Side(border_style='thin', color='FFFFFF')
    border_h_right = Side(border_style='thin', color='FFFFFF')
    border_h_left = Side(border_style='thin', color='FFFFFF')
    border_h_bottom = Side(border_style='thin', color='FFFFFF')

    border_alert_top = Side(border_style='thick', color='000000')
    border_alert_right = Side(border_style='thick', color='000000')
    border_alert_left = Side(border_style='thick', color='000000')
    border_alert_bottom = Side(border_style='thick', color='000000')

    border_ad_info_top = Side(border_style='thin', color='000000')
    border_ad_info_right = Side(border_style='thin', color='000000')
    border_ad_info_left = Side(border_style='thin', color='000000')
    border_ad_info_bottom = Side(border_style='thin', color='000000')

    alert_style = NamedStyle('alert_style')
    alert_style.font = Font('Montserrat', size=12, bold=True, color='000000')
    alert_style.fill = PatternFill('solid', fgColor="FFFFFF")
    alert_style.border = Border(top=border_alert_top, bottom=border_alert_bottom, left=border_alert_left,
                                right=border_alert_right)
    alert_style.alignment = Alignment(horizontal='center', vertical='center')

    target_style = NamedStyle('target_style')
    target_style.font = Font('Montserrat', size=10, bold=False, color='000000')
    target_style.fill = PatternFill('solid', fgColor="FFFFFF")
    target_style.border = Border(top=border_top, bottom=border_bottom, left=border_left, right=border_right)

    target_date_style = NamedStyle('target_date_style')
    target_date_style.font = Font('Montserrat', size=10, bold=False, color='000000')
    target_date_style.fill = PatternFill('solid', fgColor="FFFFFF")
    target_date_style.border = Border(top=border_top, bottom=border_bottom, left=border_left, right=border_right)
    target_date_style.number_format = 'D/MM/YYYY'

    header_style = NamedStyle('header_style')
    header_style.font = Font('Montserrat', size=11, bold=True, color='FFFFFF')
    header_style.fill = PatternFill('solid', fgColor="D21034")
    header_style.alignment = Alignment(horizontal='center', vertical='center')
    header_style.border = Border(top=border_h_top, bottom=border_h_bottom, left=border_h_left, right=border_h_right)

    ad_info_coluna_A = NamedStyle('ad_info_coluna_A')
    ad_info_coluna_A.font = Font('Montserrat', size=10, bold=True, color='FFFFFF')
    ad_info_coluna_A.fill = PatternFill('solid', fgColor="808080")
    ad_info_coluna_A.alignment = Alignment(horizontal='left', vertical='bottom')
    ad_info_coluna_A.border = Border(top=border_ad_info_top, bottom=border_ad_info_bottom,
                                     left=border_ad_info_left, right=border_ad_info_right)

    ad_info_coluna_B = NamedStyle('ad_info_coluna_B')
    ad_info_coluna_B.font = Font('Montserrat', size=10, bold=False, color='000000')
    ad_info_coluna_B.fill = PatternFill('solid', fgColor="bdbec0")
    ad_info_coluna_B.alignment = Alignment(horizontal='left', vertical='justify')
    ad_info_coluna_B.border = Border(top=border_ad_info_top, bottom=border_ad_info_bottom,
                                     left=border_ad_info_left, right=border_ad_info_right)

    relatorio_maquina = openpyxl.load_workbook(caminho)

    guias = relatorio_maquina.sheetnames

    relatorio_maquina[guias[1]].sheet_properties.tabColor = '5c5e61'
    relatorio_maquina[guias[5]].sheet_properties.tabColor = 'd21034'
    relatorio_maquina[guias[6]].sheet_properties.tabColor = '5c5e61'
    relatorio_maquina[guias[7]].sheet_properties.tabColor = '5c5e61'
    relatorio_maquina[guias[8]].sheet_properties.tabColor = 'd21034'
    relatorio_maquina[guias[9]].sheet_properties.tabColor = 'd21034'
    relatorio_maquina[guias[10]].sheet_properties.tabColor = '5c5e61'
    relatorio_maquina[guias[11]].sheet_properties.tabColor = '5c5e61'
    relatorio_maquina[guias[12]].sheet_properties.tabColor = '5c5e61'
    relatorio_maquina[guias[13]].sheet_properties.tabColor = 'd21034'

    coluna = ''

    # AD INFO ASSET GLOSSARY

    logo_a = openpyxl.drawing.image.Image(f'img/logo_swo_small.png')
    logo_b = openpyxl.drawing.image.Image(f'img/logo_swo_small.png')

    relatorio_maquina[guias[0]].add_image(logo_a, 'A1')
    relatorio_maquina[guias[0]].add_image(logo_b, 'A12')

    relatorio_maquina[guias[0]].sheet_view.showGridLines = False
    altura_linha_um = relatorio_maquina[guias[0]].row_dimensions[1]
    altura_linha_doze = relatorio_maquina[guias[0]].row_dimensions[12]
    altura_linha_um.height = 36
    altura_linha_doze.height = 36
    relatorio_maquina[guias[0]].cell(row=1, column=1).style = header_style
    relatorio_maquina[guias[0]].cell(row=1, column=2).style = header_style
    relatorio_maquina[guias[0]].cell(row=1, column=2).value = 'AD Info Asset - Glossary of Terms'
    relatorio_maquina[guias[0]].cell(row=12, column=1).style = header_style
    relatorio_maquina[guias[0]].cell(row=12, column=2).style = header_style
    relatorio_maquina[guias[0]].cell(row=12, column=2).value = 'Definition of Headers'
    relatorio_maquina[guias[0]].column_dimensions['A'].width = 45
    relatorio_maquina[guias[0]].column_dimensions['B'].width = 180

    # Inserindo texto
    cleaned = relatorio_maquina[guias[0]].cell(row=2, column=1)
    duplicate = relatorio_maquina[guias[0]].cell(row=3, column=1)
    non = relatorio_maquina[guias[0]].cell(row=4, column=1)
    unknown = relatorio_maquina[guias[0]].cell(row=5, column=1)
    dta = relatorio_maquina[guias[0]].cell(row=6, column=1)
    failover = relatorio_maquina[guias[0]].cell(row=7, column=1)
    disabled = relatorio_maquina[guias[0]].cell(row=8, column=1)
    last = relatorio_maquina[guias[0]].cell(row=9, column=1)
    raw = relatorio_maquina[guias[0]].cell(row=10, column=1)

    cleaned.value = '1. Cleaned Up AD List (In Scope)'
    duplicate.value = '2. Duplicate Computers'
    non.value = '3. NON MS-OS'
    unknown.value = '4. Unknown OS (In Scope)'
    dta.value = '5. DTA Devices (In Scope)'
    failover.value = '6. Failover Clusters'
    disabled.value = '7. Disabled'
    last.value = '8. >' + str(qnty_dias) + ' Days Last Active'
    raw.value = '9. Raw Data'

    cleaned = relatorio_maquina[guias[0]].cell(row=2, column=2)
    duplicate = relatorio_maquina[guias[0]].cell(row=3, column=2)
    non = relatorio_maquina[guias[0]].cell(row=4, column=2)
    unknown = relatorio_maquina[guias[0]].cell(row=5, column=2)
    dta = relatorio_maquina[guias[0]].cell(row=6, column=2)
    failover = relatorio_maquina[guias[0]].cell(row=7, column=2)
    disabled = relatorio_maquina[guias[0]].cell(row=8, column=2)
    last = relatorio_maquina[guias[0]].cell(row=9, column=2)
    raw = relatorio_maquina[guias[0]].cell(row=10, column=2)

    cleaned.value = "From the Raw Data, SoftwareONE scrubs the data and the devices in this tab are the " \
                    "active assets deployed in the customer's environment which are in scope within " \
                    "60 days. Anything older than 60 days is considered out of scope / decommissioned " \
                    "/ Retired (see Tab 8. >60 Days Last Active)"
    duplicate.value = 'These are devices that have reported more than once. These devices need to be ' \
                      'excluded so as to represent a true count of the actual number of unique devices.'
    non.value = 'These are devices that do not have a Microsoft Operating System (MS OS). The devices in ' \
                'scope for Microsoft engagements are typically just Microsoft Operating Systems.'
    unknown.value = 'These are devices where the Operating System has reported as Unknown or Blank.'
    dta.value = 'Developer, Test and User Acceptance (DTA) devices are listed here. Here, ' \
                'SoftwareONE has found naming conventions such as Dev, Test, Patch, TST etc. in ' \
                'the device name or DNS host name or description, and these have been classified ' \
                'these as DTA machines. '
    failover.value = "These are the devices which will never have agent on them because they " \
                     "don't exist so they require no license."
    disabled.value = 'These are devices that have been disabled by the system administrator. ' \
                     'These devices are considered as no longer running production workloads ' \
                     'and have been removed by the system administrator from the network. ' \
                     'Disabled devices within last 60 days are excluded and considered out of scope.'
    last.value = 'These are the devices that are older than 60 days and are considered out of scope ' \
                 '/ decommissioned / Retired.'
    raw.value = 'This is the Raw original data output before any analysis has been done. ' \
                'Data kept for future reference'

    cleaned = relatorio_maquina[guias[0]].cell(row=13, column=1)
    duplicate = relatorio_maquina[guias[0]].cell(row=14, column=1)
    non = relatorio_maquina[guias[0]].cell(row=15, column=1)
    unknown = relatorio_maquina[guias[0]].cell(row=16, column=1)
    dta = relatorio_maquina[guias[0]].cell(row=17, column=1)
    failover = relatorio_maquina[guias[0]].cell(row=18, column=1)
    disabled = relatorio_maquina[guias[0]].cell(row=19, column=1)
    last = relatorio_maquina[guias[0]].cell(row=20, column=1)
    raw = relatorio_maquina[guias[0]].cell(row=21, column=1)
    parent = relatorio_maquina[guias[0]].cell(row=22, column=1)

    cleaned.value = 'Source'
    duplicate.value = 'Name'
    non.value = 'Type of Device'
    unknown.value = 'Creation Date'
    dta.value = 'Description'
    failover.value = 'Disabled'
    disabled.value = 'DNS Host Name'
    last.value = 'Last Logon Date'
    raw.value = 'Operating System'
    parent.value = 'Parent Container'

    cleaned = relatorio_maquina[guias[0]].cell(row=13, column=2)
    duplicate = relatorio_maquina[guias[0]].cell(row=14, column=2)
    non = relatorio_maquina[guias[0]].cell(row=15, column=2)
    unknown = relatorio_maquina[guias[0]].cell(row=16, column=2)
    dta = relatorio_maquina[guias[0]].cell(row=17, column=2)
    failover = relatorio_maquina[guias[0]].cell(row=18, column=2)
    disabled = relatorio_maquina[guias[0]].cell(row=19, column=2)
    last = relatorio_maquina[guias[0]].cell(row=20, column=2)
    raw = relatorio_maquina[guias[0]].cell(row=21, column=2)
    parent = relatorio_maquina[guias[0]].cell(row=22, column=2)

    cleaned.value = 'Data source of the file - typically the domain name'
    duplicate.value = 'The name of the device'
    non.value = 'The Devices are categorized as Client, Server & Unknown based upon the ' \
                'Operating System & Parent Container columns.'
    unknown.value = 'The date and time that the object was originally created'
    dta.value = 'Description of the object, typically assigned by an Administrator'
    failover.value = 'Indicates whether or not this computer account has been disabled'
    disabled.value = "DNS stands for Domain Name System which is a hierarchical naming system created" \
                     " for translating host names to IP addresses. A Host Name is simply a name " \
                     "identifying a computer on a network or a domain on the Internet. "
    last.value = 'The date and time that this computer last logged on to the network. ' \
                 'Please note that if computers are only ever used remotely over a VPN ' \
                 'connection then it is possible that this value will not be an accurate' \
                 ' representation of the last time the computer was used'
    raw.value = 'The OS that is running on the computer that this object represents. ' \
                'For example: Windows XP Professional'
    parent.value = 'The full path to the container that this ' \
                   'object is a child of. Usually an OU or Container'

    # Gráficos

    for i in range(2, 6):
        num_l = int(get_numero_linhas(i, caminho))
        grafico = PieChart()
        if i == 2:
            labels = Reference(relatorio_maquina[guias[i]], min_col=1, min_row=2, max_row=3)
            data = Reference(relatorio_maquina[guias[i]], min_col=2, min_row=1, max_row=3)
            grafico.title = "Quantity of Machines"
            relatorio_maquina[guias[1]].add_chart(grafico, "D1")
        elif i == 3:
            labels = Reference(relatorio_maquina[guias[i]], min_col=1, min_row=2, max_row=3)
            data = Reference(relatorio_maquina[guias[i]], min_col=2, min_row=1, max_row=3)
            grafico.title = "In-Scope Devices"
            relatorio_maquina[guias[1]].add_chart(grafico, "M1")
        elif i == 4:
            labels = Reference(relatorio_maquina[guias[i]], min_col=1, min_row=2, max_row=3)
            data = Reference(relatorio_maquina[guias[i]], min_col=2, min_row=1, max_row=3)
            grafico.title = "Developer/Test/Acceptance & Unknown OS (In-Scope)"
            relatorio_maquina[guias[1]].add_chart(grafico, "D16")
        elif i == 5:
            labels = Reference(relatorio_maquina[guias[i]], min_col=1, min_row=2, max_row=5)
            data = Reference(relatorio_maquina[guias[i]], min_col=2, min_row=1, max_row=5)
            grafico.title = "Types of Out of Scope"
            relatorio_maquina[guias[1]].add_chart(grafico, "M16")

        grafico.add_data(data, titles_from_data=True)
        grafico.show_number = True
        grafico.set_categories(labels)

    for linha in range(2, 11):
        relatorio_maquina[guias[0]].cell(row=linha, column=1).style = ad_info_coluna_A
        relatorio_maquina[guias[0]].cell(row=linha, column=2).style = ad_info_coluna_B

    for linha in range(13, 23):
        relatorio_maquina[guias[0]].cell(row=linha, column=1).style = ad_info_coluna_A
        relatorio_maquina[guias[0]].cell(row=linha, column=2).style = ad_info_coluna_B

    # Gráficos
    relatorio_maquina[guias[1]].sheet_view.showGridLines = False

    for i in range(2, 6):
        relatorio_maquina[guias[i]].sheet_state = 'hidden'

    for i in range(6, 15):
        numero_colunas = get_numero_colunas(i, caminho)
        numero_linhas = get_numero_linhas(i, caminho)
        verify = relatorio_maquina[guias[i]]['A2'].value
        for j in range(1, numero_colunas + 1):
            if j == 1:
                coluna = 'A'
            if j == 2:
                coluna = 'B'
            if j == 3:
                coluna = 'C'
            if j == 4:
                coluna = 'D'
            if j == 5:
                coluna = 'E'
            if j == 6:
                coluna = 'F'
            if j == 7:
                coluna = 'G'
            if j == 8:
                coluna = 'H'
            for k in range(1, numero_linhas + 2):
                if verify is None:
                    relatorio_maquina[guias[i]].merge_cells('A2:C5')
                    celula = relatorio_maquina[guias[i]].cell(row=2, column=1)
                    celula.value = 'NO DATA FOUND'
                    celula.style = alert_style
                if (j == 3) | (j == 6):
                    relatorio_maquina[guias[i]].cell(row=k, column=j).style = target_date_style
                else:
                    relatorio_maquina[guias[i]].cell(row=k, column=j).style = target_style

            relatorio_maquina[guias[i]].cell(row=1, column=j).style = header_style
            relatorio_maquina[guias[i]].column_dimensions[coluna].width = 40

        relatorio_maquina[guias[i]].auto_filter.ref = relatorio_maquina[guias[i]].dimensions
        relatorio_maquina[guias[i]].sheet_view.showGridLines = False
        altura = relatorio_maquina[guias[i]].row_dimensions[1]
        altura.height = 36

    relatorio_maquina.save(caminho)
